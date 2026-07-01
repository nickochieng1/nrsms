from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.dependencies import get_audit_meta, get_current_user, require_role
from app.crud import submission as crud_sub
from app.db.session import get_db
from app.models.submission import Submission, SubmissionStatus
from app.models.user import User, UserRole
from app.schemas.submission import (
    SubmissionCreate, SubmissionOut, SubmissionRegionalStatusRow, SubmissionReview, SubmissionUpdate,
)
from app.services import audit as audit_svc
from app.services.validation import validate_submission

router = APIRouter(prefix="/submissions", tags=["submissions"])


# Statuses each downstream role can see once a submission has left DRAFT —
# narrower than "everything" so e.g. a CROP in Mombasa never sees Kisumu's
# queue, but broad enough that everyone can see the history of something
# they already acted on.
NON_DRAFT = [s for s in SubmissionStatus if s != SubmissionStatus.DRAFT]
HQ_CLERK_VISIBLE = [SubmissionStatus.RROP_APPROVED, SubmissionStatus.HQ_COMPILED, SubmissionStatus.APPROVED]
REGISTRAR_VISIBLE = [SubmissionStatus.HQ_COMPILED, SubmissionStatus.APPROVED]
DIRECTOR_VISIBLE = [SubmissionStatus.APPROVED]


@router.get("", response_model=List[SubmissionOut])
def list_submissions(
    status: Optional[SubmissionStatus] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    skip: int = 0,
    limit: int = 200,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    role = current_user.role

    if role in (UserRole.DCROP, UserRole.CLERK):
        # Legacy CLERK accounts behave like a DCROP scoped to their region
        # (no subcounty/county filter) until reassigned to the new role.
        return crud_sub.get_all(
            db, subcounty=current_user.subcounty, county=current_user.county, region=current_user.region,
            submitted_by=current_user.id, status=status, year=year, month=month, skip=skip, limit=limit,
        )

    if role == UserRole.CROP:
        if not current_user.county:
            return []
        return crud_sub.get_all(
            db, county=current_user.county, statuses=NON_DRAFT,
            status=status, year=year, month=month, skip=skip, limit=limit,
        )

    if role == UserRole.RROP:
        if not current_user.region:
            return []
        return crud_sub.get_all(
            db, region=current_user.region, statuses=NON_DRAFT,
            status=status, year=year, month=month, skip=skip, limit=limit,
        )

    if role == UserRole.HQ_CLERK:
        if not current_user.region:
            return []
        return crud_sub.get_all(
            db, region=current_user.region, statuses=HQ_CLERK_VISIBLE,
            status=status if status in HQ_CLERK_VISIBLE else None, year=year, month=month, skip=skip, limit=limit,
        )

    if role == UserRole.REGISTRAR:
        return crud_sub.get_all(
            db, statuses=REGISTRAR_VISIBLE,
            status=status if status in REGISTRAR_VISIBLE else None, year=year, month=month, skip=skip, limit=limit,
        )

    if role == UserRole.DIRECTOR:
        return crud_sub.get_all(
            db, statuses=DIRECTOR_VISIBLE,
            status=status if status in DIRECTOR_VISIBLE else None, year=year, month=month, skip=skip, limit=limit,
        )

    return []


@router.post("", response_model=SubmissionOut, status_code=status.HTTP_201_CREATED)
def create_submission(
    body: SubmissionCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.DCROP, UserRole.CLERK)),
):
    if not current_user.subcounty or not current_user.county or not current_user.region:
        raise HTTPException(
            status_code=403,
            detail="Your account has no subcounty/county/region assigned — ask an admin to set it",
        )

    warnings = validate_submission(body)
    submission = crud_sub.create(
        db, body, current_user.id,
        subcounty=current_user.subcounty, county=current_user.county, region=current_user.region,
    )
    meta = get_audit_meta(request)
    audit_svc.log(
        db, user_id=current_user.id, action="CREATE", resource="submission",
        resource_id=submission.id,
        new_value={"subcounty": current_user.subcounty, "county": current_user.county,
                   "period": f"{body.period_month}/{body.period_year}", "warnings": warnings},
        **meta,
    )
    return submission


@router.get("/regional-status", response_model=List[SubmissionRegionalStatusRow])
def regional_status(
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(
        UserRole.RROP, UserRole.HQ_CLERK, UserRole.REGISTRAR, UserRole.DIRECTOR, UserRole.ADMIN,
    )),
):
    """Per-region breakdown of where each region's submissions sit in the
    approval chain for one period — powers the Dashboard pie/bar charts."""
    rows = (
        db.query(Submission.region, Submission.status, func.count(Submission.id))
        .filter(Submission.period_year == year, Submission.period_month == month, Submission.region.isnot(None))
        .group_by(Submission.region, Submission.status)
        .all()
    )
    by_region: dict = {}
    for region, st, count in rows:
        row = by_region.setdefault(region, {
            "region": region, "not_started": 0, "dcrop_submitted": 0, "crop_approved": 0,
            "rrop_approved": 0, "hq_compiled": 0, "approved": 0, "total": 0,
        })
        key = st.value if hasattr(st, "value") else st
        if key in row:
            row[key] += count
        row["total"] += count
    return sorted(by_region.values(), key=lambda r: r["region"])


# ── Excel bulk upload ─────────────────────────────────────────────────────────
import io as _io
import openpyxl as _openpyxl
from fastapi import UploadFile, File


BULK_FIELDS = [
    "period_month", "period_year",
    "app_npr_male", "app_npr_female",
    "app_replacements_male", "app_replacements_female",
    "app_changes_male", "app_changes_female",
    "app_duplicates_male", "app_duplicates_female",
    "app_type4_male", "app_type4_female",
    "app_type5_male", "app_type5_female",
    "ids_npr_male", "ids_npr_female",
    "ids_replacements_male", "ids_replacements_female",
    "ids_changes_male", "ids_changes_female",
    "ids_duplicates_male", "ids_duplicates_female",
    "ids_type4_male", "ids_type4_female",
    "ids_type5_male", "ids_type5_female",
    "rej_npr_male", "rej_npr_female",
    "rej_replacements_male", "rej_replacements_female",
    "rej_changes_male", "rej_changes_female",
    "rej_duplicates_male", "rej_duplicates_female",
    "rej_type4_male", "rej_type4_female",
    "rej_type5_male", "rej_type5_female",
    "collected_npr_male", "collected_npr_female",
    "collected_others_male", "collected_others_female",
    "collected_rejected_male", "collected_rejected_female",
    "uncollected_npr_male", "uncollected_npr_female",
    "uncollected_others_male", "uncollected_others_female",
    "uncollected_lost_male", "uncollected_lost_female",
    "reg136c_balance_bd", "reg136c_used", "reg136c_spoilt", "reg136c_returned",
    "photo3a_balance_bd", "photo3a_used", "photo3a_spoilt", "photo3a_returned",
]


@router.get("/bulk-template")
def download_bulk_template():
    """Return an Excel template the DCROP can fill in and re-upload."""
    from fastapi.responses import Response
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submissions"
    ws.append(BULK_FIELDS)
    ws.append([6, 2026] + [0] * (len(BULK_FIELDS) - 2))  # example row
    buf = _io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=nrsms_bulk_template.xlsx"},
    )


@router.post("/bulk-upload")
def bulk_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.DCROP, UserRole.CLERK)),
):
    """Parse a filled-in Excel template and create a submission for each row.
    Returns a list of {row, status, detail} objects."""
    if not current_user.subcounty and not current_user.county:
        raise HTTPException(status_code=403, detail="Your account has no geographic area assigned")

    try:
        wb = _openpyxl.load_workbook(_io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse the uploaded file — must be .xlsx")

    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "period_month" not in headers:
        raise HTTPException(status_code=400, detail="First row must be headers matching the template")

    results = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None for v in row):
            continue
        row_data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        try:
            payload = SubmissionCreate(
                period_month=int(row_data.get("period_month") or 0),
                period_year=int(row_data.get("period_year") or 0),
                **{f: int(row_data.get(f) or 0) for f in BULK_FIELDS if f not in ("period_month", "period_year")},
            )
            sub = crud_sub.create(
                db, payload, current_user.id,
                subcounty=current_user.subcounty, county=current_user.county, region=current_user.region,
            )
            results.append({"row": row_idx, "status": "created", "id": sub.id,
                            "detail": f"{payload.period_month}/{payload.period_year}"})
        except Exception as e:
            results.append({"row": row_idx, "status": "error", "detail": str(e)})

    return {"results": results, "created": sum(1 for r in results if r["status"] == "created")}

@router.get("/{submission_id}", response_model=SubmissionOut)
def get_submission(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    _assert_can_read(current_user, sub)
    return sub


@router.patch("/{submission_id}", response_model=SubmissionOut)
def update_submission(
    submission_id: int,
    body: SubmissionUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    if sub.status not in (SubmissionStatus.DRAFT, SubmissionStatus.CROP_REJECTED):
        raise HTTPException(status_code=400, detail="Only draft or CROP-rejected submissions can be edited")
    if current_user.role in (UserRole.DCROP, UserRole.CLERK) and sub.submitted_by != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    updated = crud_sub.update(db, sub, body)
    meta = get_audit_meta(request)
    audit_svc.log(db, user_id=current_user.id, action="UPDATE", resource="submission",
                  resource_id=submission_id,
                  new_value=body.model_dump(exclude_unset=True), **meta)
    return updated


@router.post("/{submission_id}/submit", response_model=SubmissionOut)
def submit_submission(
    submission_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.DCROP, UserRole.CLERK)),
):
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    if sub.submitted_by != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    if sub.status not in (SubmissionStatus.DRAFT, SubmissionStatus.CROP_REJECTED):
        raise HTTPException(status_code=400, detail="Only draft or CROP-rejected submissions can be submitted")
    updated = crud_sub.submit(db, sub)
    meta = get_audit_meta(request)
    audit_svc.log(db, user_id=current_user.id, action="SUBMIT", resource="submission",
                  resource_id=submission_id, **meta)
    # Notify CROP(s) in the same county that data awaits review
    period = f"{updated.period_month}/{updated.period_year}"
    return updated


@router.post("/{submission_id}/crop-review", response_model=SubmissionOut)
def crop_review_submission(
    submission_id: int,
    body: SubmissionReview,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.CROP)),
):
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    if current_user.county and sub.county and sub.county.lower() != current_user.county.lower():
        raise HTTPException(status_code=403, detail="Access denied")
    if sub.status not in (SubmissionStatus.DCROP_SUBMITTED, SubmissionStatus.RROP_REJECTED):
        raise HTTPException(status_code=400, detail="This submission is not awaiting CROP review")

    meta = get_audit_meta(request)
    period = f"{sub.period_month}/{sub.period_year}"
    if body.action == "approve":
        updated = crud_sub.crop_approve(db, sub, current_user.id)
        audit_svc.log(db, user_id=current_user.id, action="CROP_APPROVE", resource="submission",
                      resource_id=submission_id, **meta)
    elif body.action == "reject":
        if not body.rejection_reason:
            raise HTTPException(status_code=400, detail="rejection_reason is required")
        updated = crud_sub.crop_reject(db, sub, current_user.id, body.rejection_reason)
        audit_svc.log(db, user_id=current_user.id, action="CROP_REJECT", resource="submission",
                      resource_id=submission_id, new_value={"reason": body.rejection_reason}, **meta)
    else:
        raise HTTPException(status_code=400, detail="action must be 'approve' or 'reject'")
    return updated


@router.post("/{submission_id}/rrop-review", response_model=SubmissionOut)
def rrop_review_submission(
    submission_id: int,
    body: SubmissionReview,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.RROP)),
):
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    if current_user.region and sub.region and sub.region.lower() != current_user.region.lower():
        raise HTTPException(status_code=403, detail="Access denied")
    if sub.status != SubmissionStatus.CROP_APPROVED:
        raise HTTPException(status_code=400, detail="This submission is not awaiting RROP review")

    meta = get_audit_meta(request)
    period = f"{sub.period_month}/{sub.period_year}"
    if body.action == "approve":
        updated = crud_sub.rrop_approve(db, sub, current_user.id)
        audit_svc.log(db, user_id=current_user.id, action="RROP_APPROVE", resource="submission",
                      resource_id=submission_id, **meta)
    elif body.action == "reject":
        if not body.rejection_reason:
            raise HTTPException(status_code=400, detail="rejection_reason is required")
        updated = crud_sub.rrop_reject(db, sub, current_user.id, body.rejection_reason)
        audit_svc.log(db, user_id=current_user.id, action="RROP_REJECT", resource="submission",
                      resource_id=submission_id, new_value={"reason": body.rejection_reason}, **meta)
    else:
        raise HTTPException(status_code=400, detail="action must be 'approve' or 'reject'")
    return updated


@router.post("/{submission_id}/hq-compile", response_model=SubmissionOut)
def hq_compile_submission(
    submission_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.HQ_CLERK)),
):
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    if current_user.region and sub.region and sub.region.lower() != current_user.region.lower():
        raise HTTPException(status_code=403, detail="Access denied")
    if sub.status != SubmissionStatus.RROP_APPROVED:
        raise HTTPException(status_code=400, detail="This submission is not awaiting HQ compilation")

    updated = crud_sub.hq_compile(db, sub, current_user.id)
    meta = get_audit_meta(request)
    audit_svc.log(db, user_id=current_user.id, action="HQ_COMPILE", resource="submission",
                  resource_id=submission_id, **meta)
    return updated


@router.post("/{submission_id}/review", response_model=SubmissionOut)
def review_submission(
    submission_id: int,
    body: SubmissionReview,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.REGISTRAR)),
):
    """Registrar's final approval — the last gate before the Director sees it."""
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    if sub.status != SubmissionStatus.HQ_COMPILED:
        raise HTTPException(status_code=400, detail="This submission is not awaiting registrar review")

    meta = get_audit_meta(request)
    period = f"{sub.period_month}/{sub.period_year}"
    if body.action == "approve":
        updated = crud_sub.approve(db, sub, current_user.id)
        audit_svc.log(db, user_id=current_user.id, action="APPROVE", resource="submission",
                      resource_id=submission_id, **meta)
    elif body.action == "reject":
        if not body.rejection_reason:
            raise HTTPException(status_code=400, detail="rejection_reason is required")
        updated = crud_sub.reject(db, sub, current_user.id, body.rejection_reason)
        audit_svc.log(db, user_id=current_user.id, action="REJECT", resource="submission",
                      resource_id=submission_id, new_value={"reason": body.rejection_reason}, **meta)
    else:
        raise HTTPException(status_code=400, detail="action must be 'approve' or 'reject'")
    return updated


def _assert_can_read(user: User, sub: Submission) -> None:
    role = user.role
    same = lambda a, b: bool(a and b and a.lower() == b.lower())  # noqa: E731

    if role == UserRole.ADMIN:
        return  # admin has emergency read access to all submissions
    if role in (UserRole.DCROP, UserRole.CLERK):
        if sub.submitted_by != user.id:
            raise HTTPException(status_code=403, detail="Access denied")
    elif role == UserRole.CROP:
        if sub.status == SubmissionStatus.DRAFT or not same(sub.county, user.county):
            raise HTTPException(status_code=403, detail="Access denied")
    elif role == UserRole.RROP:
        if sub.status == SubmissionStatus.DRAFT or not same(sub.region, user.region):
            raise HTTPException(status_code=403, detail="Access denied")
    elif role == UserRole.HQ_CLERK:
        if sub.status not in HQ_CLERK_VISIBLE or not same(sub.region, user.region):
            raise HTTPException(status_code=403, detail="Access denied")
    elif role == UserRole.REGISTRAR:
        if sub.status not in REGISTRAR_VISIBLE:
            raise HTTPException(status_code=403, detail="Access denied")
    elif role == UserRole.DIRECTOR:
        if sub.status not in DIRECTOR_VISIBLE:
            raise HTTPException(status_code=403, detail="Access denied")
    else:
        raise HTTPException(status_code=403, detail="Access denied")


# ── Comments ──────────────────────────────────────────────────────────────────
from typing import List as TypingList

from app.models.submission_comment import SubmissionComment
from app.schemas.submission_comment import CommentCreate, CommentOut


@router.get("/{submission_id}/comments", response_model=TypingList[CommentOut])
def list_comments(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    _assert_can_read(current_user, sub)
    return db.query(SubmissionComment).filter(
        SubmissionComment.submission_id == submission_id
    ).order_by(SubmissionComment.created_at).all()


@router.post("/{submission_id}/comments", response_model=CommentOut, status_code=201)
def add_comment(
    submission_id: int,
    body: CommentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    _assert_can_read(current_user, sub)
    if not body.content.strip():
        raise HTTPException(status_code=400, detail="Comment cannot be empty")
    comment = SubmissionComment(
        submission_id=submission_id,
        user_id=current_user.id,
        content=body.content.strip(),
    )
    db.add(comment)
    db.commit()
    db.refresh(comment)
    return comment


# ── Anomaly check ─────────────────────────────────────────────────────────────
from app.services.anomaly import check_anomalies


@router.get("/{submission_id}/anomaly-check")
def anomaly_check(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return a list of warning strings if any field is anomalously high
    vs the submitter's own 6-month history. Empty list = all looks normal."""
    sub = crud_sub.get(db, submission_id)
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    _assert_can_read(current_user, sub)
    return {"warnings": check_anomalies(db, sub)}


