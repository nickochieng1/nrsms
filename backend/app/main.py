from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware

from app.api.v1.router import api_router
from app.core.config import settings
from app.db.base import Base
from app.db.session import engine

app = FastAPI(
    title="NRSMS API",
    description="National Registration Statistics Management System",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:3000",
        "tauri://localhost",
        "https://tauri.localhost",
    ],
    allow_origin_regex=r"https://.*\.netlify\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(api_router)


@app.on_event("startup")
def on_startup():
    Base.metadata.create_all(bind=engine)
    _migrate_schema()
    _migrate_audit_log_schema()
    _seed_superuser()
    _seed_stations()


def _seed_superuser():
    from sqlalchemy.orm import Session
    from app.core.security import get_password_hash
    from app.crud import user as crud_user
    from app.models.user import UserRole
    from app.schemas.user import UserCreate

    with Session(engine) as db:
        existing = crud_user.get_by_email(db, settings.FIRST_SUPERUSER_EMAIL)
        if not existing:
            crud_user.create(db, UserCreate(
                full_name="System Administrator",
                username="Administrator",
                email=settings.FIRST_SUPERUSER_EMAIL,
                password=settings.FIRST_SUPERUSER_PASSWORD,
                role=UserRole.ADMIN,
            ))
        elif existing.username != "Administrator":
            # One-time fixup: the superuser was originally seeded as "admin"
            # with whatever FIRST_SUPERUSER_PASSWORD was at the time — bumping
            # that env var later never touched the already-created account.
            # Move it to the username/password NRB actually wants once; once
            # the username matches, this is a no-op on every later deploy, so
            # it won't keep clobbering a password someone changes afterward.
            taken_by = crud_user.get_by_username(db, "Administrator")
            if taken_by is None:
                existing.username = "Administrator"
                existing.hashed_password = get_password_hash(settings.FIRST_SUPERUSER_PASSWORD)
                db.commit()


def _migrate_schema():
    """Add new columns and remap old role/status values to new ones.

    This whole function is SQLite-only patchwork accumulated over the app's
    history (ALTER TABLE/PRAGMA workarounds for things SQLite can't do
    directly). A fresh Postgres database gets the current schema straight
    from the models via create_all() above, so none of this applies there.
    """
    if not settings.DATABASE_URL.startswith("sqlite"):
        return
    from sqlalchemy import text
    with engine.connect() as conn:
        for col in ("county VARCHAR(200)", "region VARCHAR(200)"):
            try:
                conn.execute(text(f"ALTER TABLE users ADD COLUMN {col}"))
                conn.commit()
            except Exception:
                pass

        try:
            conn.execute(text("ALTER TABLE mobile_registrations ADD COLUMN target_set INTEGER DEFAULT 0"))
            conn.commit()
        except Exception:
            pass

        for col in (
            "is_closed BOOLEAN DEFAULT 0",
            "closed_at DATETIME",
            "closed_by INTEGER",
            "created_by INTEGER",
        ):
            try:
                conn.execute(text(f"ALTER TABLE mobile_registrations ADD COLUMN {col}"))
                conn.commit()
            except Exception:
                pass

        try:
            conn.execute(text("UPDATE mobile_registrations SET created_by=submitted_by WHERE created_by IS NULL"))
            conn.execute(text("UPDATE mobile_registrations SET is_closed=1 WHERE lower(status)='approved'"))
            conn.commit()
        except Exception:
            pass

        # status/submitted_by are NOT NULL legacy columns with no default, so
        # once the model stopped setting them, every new insert started
        # failing. `status` can be dropped directly; `submitted_by` carries a
        # FOREIGN KEY, which SQLite's DROP COLUMN refuses to touch, so that
        # one needs a rename-recreate-copy (data already migrated onto
        # created_by above).
        for col in ("status",):
            try:
                conn.execute(text(f"ALTER TABLE mobile_registrations DROP COLUMN {col}"))
                conn.commit()
            except Exception:
                pass

        try:
            cols = [r[1] for r in conn.execute(text("PRAGMA table_info(mobile_registrations)")).fetchall()]
            if "submitted_by" in cols:
                new_cols = [
                    "id", "county", "subcounty", "period_month", "period_year", "notes", "created_by",
                    "is_closed", "closed_at", "closed_by",
                    "age_25_40_male", "age_25_40_female", "age_25_40_total",
                    "age_41_60_male", "age_41_60_female", "age_41_60_total",
                    "age_60_plus_male", "age_60_plus_female", "age_60_plus_total",
                    "created_at", "updated_at",
                ]
                col_csv = ", ".join(new_cols)
                conn.execute(text("ALTER TABLE mobile_registrations RENAME TO mobile_registrations_legacy"))
                # SQLite index names are global, not per-table — renaming the table
                # keeps the old index names attached, which collide with the new
                # table's indexes below. Drop them; the legacy table is dropped anyway.
                conn.execute(text("DROP INDEX IF EXISTS ix_mobile_registrations_id"))
                conn.execute(text("DROP INDEX IF EXISTS ix_mobile_registrations_county"))
                conn.commit()
                Base.metadata.tables["mobile_registrations"].create(bind=engine)
                conn.execute(text(
                    f"INSERT INTO mobile_registrations ({col_csv}) SELECT {col_csv} FROM mobile_registrations_legacy"
                ))
                conn.execute(text("DROP TABLE mobile_registrations_legacy"))
                conn.commit()
        except Exception:
            conn.rollback()

        # Target moved from per-subcounty-exercise to per-county
        # (MobileRegistrationTarget, created automatically by create_all above).
        # Backfill any existing target_set values, deduping by county+period,
        # before dropping the now-unused column.
        try:
            cols = [r[1] for r in conn.execute(text("PRAGMA table_info(mobile_registrations)")).fetchall()]
            if "target_set" in cols:
                conn.execute(text("""
                    INSERT OR IGNORE INTO mobile_registration_targets
                        (county, period_month, period_year, target_set, set_by, created_at, updated_at)
                    SELECT county, period_month, period_year, MAX(target_set), NULL,
                           CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
                    FROM mobile_registrations
                    WHERE target_set > 0
                    GROUP BY county, period_month, period_year
                """))
                conn.commit()
                conn.execute(text("ALTER TABLE mobile_registrations DROP COLUMN target_set"))
                conn.commit()
        except Exception:
            conn.rollback()

        role_map = [
            ("station_officer", "clerk"),
            ("registrar", "registrar"),
            ("sub_county_registrar", "registrar"),
            ("county_registrar", "registrar"),
            ("regional_registrar", "registrar"),
            ("hq_clerk", "director"),
            ("hq_officer", "director"),
        ]
        for old, new in role_map:
            conn.execute(text("UPDATE users SET role=:n WHERE role=:o"), {"n": new, "o": old})
        conn.commit()

        status_map = [
            ("under_review", "submitted"),
            ("registrar_approved", "approved"),
            ("sub_county_approved", "approved"),
            ("county_approved", "approved"),
            ("regional_approved", "approved"),
        ]
        for old, new in status_map:
            conn.execute(text("UPDATE submissions SET status=:n WHERE status=:o"), {"n": new, "o": old})
        conn.commit()

        # Copy station region onto clerks that only had station_id
        conn.execute(text("""
            UPDATE users
            SET region = (
                SELECT region FROM stations WHERE stations.id = users.station_id
            )
            WHERE role = 'clerk'
              AND region IS NULL
              AND station_id IS NOT NULL
        """))
        conn.commit()


def _migrate_audit_log_schema():
    """Add the actor-identity snapshot columns to audit_logs and backfill
    them for existing rows. Unlike _migrate_schema() above, this runs on
    every database (Postgres included) — audit_logs already existed in
    production before these columns were added, so create_all() alone
    won't retrofit them onto the live table.
    """
    from sqlalchemy import text
    is_sqlite = settings.DATABASE_URL.startswith("sqlite")
    with engine.connect() as conn:
        if is_sqlite:
            for col in ("actor_name VARCHAR(200)", "actor_username VARCHAR(100)", "actor_role VARCHAR(50)"):
                try:
                    conn.execute(text(f"ALTER TABLE audit_logs ADD COLUMN {col}"))
                    conn.commit()
                except Exception:
                    pass
        else:
            for col in ("actor_name VARCHAR(200)", "actor_username VARCHAR(100)", "actor_role VARCHAR(50)"):
                conn.execute(text(f"ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS {col}"))
            conn.commit()

        # Best-effort backfill from the live users table — only recovers
        # identity for entries whose user_id FK hasn't already been nulled
        # out (e.g. by a since-deleted account); anything already nulled
        # has no name to recover and stays "Unknown user" in the UI.
        # LOWER() on the role: Postgres' native enum column stores the
        # member NAME ("ADMIN"), but the app snapshots .value ("admin")
        # going forward — normalize so both forms match the same casing.
        # (Postgres needs an explicit ::text cast — LOWER() doesn't accept
        # a native enum type directly; SQLite's role column is plain text.)
        role_expr = "LOWER(role::text)" if not is_sqlite else "LOWER(role)"
        conn.execute(text(f"""
            UPDATE audit_logs
            SET actor_name = (SELECT full_name FROM users WHERE users.id = audit_logs.user_id),
                actor_username = (SELECT username FROM users WHERE users.id = audit_logs.user_id),
                actor_role = (SELECT {role_expr} FROM users WHERE users.id = audit_logs.user_id)
            WHERE user_id IS NOT NULL AND actor_name IS NULL
        """))
        conn.execute(text("""
            UPDATE audit_logs SET actor_role = LOWER(actor_role)
            WHERE actor_role IS NOT NULL AND actor_role <> LOWER(actor_role)
        """))
        conn.commit()


def _seed_stations():
    import os
    from sqlalchemy.orm import Session
    from app.models.station import Station

    xlsx_path = os.path.join(os.path.dirname(__file__), "..", "stats.xlsx")
    if not os.path.exists(xlsx_path):
        return

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    # Use the applications sheet which has the full station hierarchy
    ws = wb["APPLICATIONS SENT"] if "APPLICATIONS SENT" in wb.sheetnames else wb.active
    region = county = ""
    skip_fragments = (
        "TOTAL", "GRAND", "LIST OF", "APPLICATION", "JANUARY", "FEBRUARY",
        "MARCH", "APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPT",
        "OCTOBER", "NOVEMBER", "DECEMBER", "NPR", "REPLACEMENTS",
    )
    stations_data = []
    for row in ws.iter_rows(values_only=True):
        val = row[0]
        if not isinstance(val, str):
            continue
        val = val.strip().upper()
        if not val or val.startswith("="):
            continue
        if "REGION" in val and "COUNTY" not in val:
            region = val.replace(" TOTALS", "").replace(" TOTAL", "").strip()
        elif "COUNTY" in val:
            county = val.replace(" COUNTY", "").strip()
        elif any(frag in val for frag in skip_fragments):
            continue
        elif region and county:
            stations_data.append((region.title(), county.title(), val.title()))

    with Session(engine) as db:
        if db.query(Station).count() > 0:
            return
        seq = 1
        for region_name, county_name, name in stations_data:
            r = "".join(w[0] for w in region_name.split() if w)[:2].upper()
            c = "".join(w[0] for w in county_name.split() if w)[:2].upper()
            code = f"{r}{c}{seq:03d}"
            seq += 1
            db.add(Station(name=name, region=region_name, county=county_name, code=code))
        db.commit()


@app.get("/health")
def health():
    return {"status": "ok"}
